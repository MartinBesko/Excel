{
 "cells": [
  {
   "cell_type": "code",
   "execution_count": 2,
   "metadata": {},
   "outputs": [],
   "source": [
    "import openpyxl as xl\n",
    "from openpyxl.chart import BarChart, Reference\n",
    "\n",
    "\n",
    "def process_workbook(filename):\n",
    "\n",
    "    wb = xl.load_workbook(filename)\n",
    "    sheet = wb[\"Sheet1\"]\n",
    "\n",
    "\n",
    "    for row in range(2, sheet.max_row + 1):\n",
    "        cell = sheet.cell(row, 3)\n",
    "        corrected_price= cell.value * 0.9\n",
    "        corrected_price_cell= sheet.cell(row, 4)\n",
    "        corrected_price_cell.value = corrected_price\n",
    "\n",
    "    values= Reference(sheet, \n",
    "    min_row=2, \n",
    "    max_row= sheet.max_row,\n",
    "    min_col=4,\n",
    "    max_col=4)\n",
    "\n",
    "    chart= BarChart()\n",
    "    chart.add_data(values)\n",
    "    sheet.add_chart(chart, 'e2')\n",
    "\n",
    "\n",
    "    wb.save(filename)"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "metadata": {},
   "outputs": [],
   "source": []
  }
 ],
 "metadata": {
  "interpreter": {
   "hash": "d773c60b364924b0641ffa12f565cf6d1ed3cb8df687750f9dad8ca064018266"
  },
  "kernelspec": {
   "display_name": "Python 3.10.0 64-bit",
   "language": "python",
   "name": "python3"
  },
  "language_info": {
   "codemirror_mode": {
    "name": "ipython",
    "version": 3
   },
   "file_extension": ".py",
   "mimetype": "text/x-python",
   "name": "python",
   "nbconvert_exporter": "python",
   "pygments_lexer": "ipython3",
   "version": "3.10.0"
  },
  "orig_nbformat": 4
 },
 "nbformat": 4,
 "nbformat_minor": 2
}
